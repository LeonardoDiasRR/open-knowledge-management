import contextlib
import io
import json
import os
import re
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook

from scripts.tabular import (
    Dataset,
    FormulaDefinition,
    load_tabular_file,
    persist_source,
    query_read_only,
    rebuild_source,
    sanitize_identifier,
)


class TabularTestCase(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.raw = self.root / "raw"
        self.raw.mkdir()
        self.db = self.root / "database" / "data.duckdb"
        self.previous_root = os.environ.get("OKM_WIKI_ROOT")
        os.environ["OKM_WIKI_ROOT"] = str(self.root)

    def tearDown(self):
        if self.previous_root is None:
            os.environ.pop("OKM_WIKI_ROOT", None)
        else:
            os.environ["OKM_WIKI_ROOT"] = self.previous_root
        self.temporary_directory.cleanup()

    def _source(self, name, content="content"):
        path = self.raw / name
        path.write_text(content, encoding="utf-8")
        return path

    def _dataset(self, name, values, **metadata):
        return Dataset(
            name=name,
            columns=["Value"],
            rows=[{"Value": value} for value in values],
            **metadata,
        )

    def test_sanitize_identifier_and_collisions(self):
        self.assertEqual(sanitize_identifier("Área 1 (Notas)"), "area_1_notas")
        self.assertEqual(sanitize_identifier(""), "unnamed")

        source = self._source("collision.csv")
        manifests = persist_source(
            self.db,
            "collision",
            source,
            "wiki/sources/collision.md",
            [
                Dataset(
                    name="Collision",
                    columns=["A B", "A-B"],
                    rows=[{"A B": 1, "A-B": 2}],
                )
            ],
        )

        self.assertEqual(
            manifests[0].columns,
            [
                ("a_b", "BIGINT"),
                ("a_b_2", "BIGINT"),
                ("source_id", "VARCHAR"),
                ("source_page", "BIGINT"),
                ("source_section", "VARCHAR"),
                ("source_sheet", "VARCHAR"),
            ],
        )

    def test_load_csv_preserves_rows_and_empty_values(self):
        path = self._source("scores.csv", "Name,Code,Value\nAna,0012,\n")

        datasets = load_tabular_file(path, "scores")

        self.assertEqual(datasets[0].columns, ["Name", "Code", "Value"])
        self.assertEqual(datasets[0].rows[0]["Code"], "0012")
        self.assertIsNone(datasets[0].rows[0]["Value"])

    def test_load_tsv_uses_tab_delimiter(self):
        path = self._source("scores.tsv", "Name\tCode\tValue\nAna\t0012\t10\n")

        datasets = load_tabular_file(path, "scores")

        self.assertEqual(datasets[0].columns, ["Name", "Code", "Value"])
        self.assertEqual(
            datasets[0].rows,
            [{"Name": "Ana", "Code": "0012", "Value": "10"}],
        )

    def test_load_tabular_file_rejects_path_outside_configured_raw(self):
        with tempfile.TemporaryDirectory() as outside_directory:
            path = Path(outside_directory) / "outside.csv"
            path.write_text("Name\nAna\n", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "configured wiki raw directory"):
                load_tabular_file(path, "outside")
            with self.assertRaisesRegex(ValueError, "configured wiki raw directory"):
                persist_source(
                    self.db,
                    "outside",
                    path,
                    "outside.md",
                    [self._dataset("Items", ["Ana"])],
                )

    def test_load_xlsx_handles_multiple_worksheets_and_formula_definitions(self):
        path = self.raw / "workbook.xlsx"
        workbook = Workbook()
        results = workbook.active
        results.title = "Results"
        results.append(["A", "B", "Total"])
        results.append([None, None, "=SUM(A2:B2)"])
        results.append([2, 3, "=SUM(A3:B3)"])
        other = workbook.create_sheet("Other")
        other.append(["Code", "Name"])
        other.append(["0012", "Ana"])
        formula_only = workbook.create_sheet("Formula Only")
        formula_only.append(["Value", "Total"])
        formula_only.append([None, "=SUM(A2:A2)"])
        workbook.create_sheet("Empty")
        workbook.save(path)
        workbook.close()

        datasets = load_tabular_file(path, "book")

        self.assertEqual([dataset.source_sheet for dataset in datasets], ["Results", "Other"])
        self.assertNotIn("Empty", [dataset.source_sheet for dataset in datasets])
        self.assertEqual(datasets[0].columns, ["A", "B"])
        self.assertEqual(datasets[0].rows, [{"A": 2, "B": 3}])
        self.assertNotIn("Total", datasets[0].columns)
        self.assertNotIn("Total", datasets[0].rows[0])
        self.assertEqual(datasets[0].formulas[0].formula, "=SUM(A2:B2)")
        self.assertIsInstance(datasets[0].formulas[0], FormulaDefinition)
        self.assertEqual(datasets[0].formulas[0].inputs, ("A", "B"))
        self.assertIsNotNone(datasets[0].formulas[0].sql)
        self.assertEqual(datasets[1].rows[0]["Code"], "0012")

    def test_cached_xlsx_formula_value_is_absent_after_persistence(self):
        path = self.raw / "cached.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["A", "B", "Total"])
        worksheet.append([2, 3, "=SUM(A2:B2)"])
        workbook.save(path)
        workbook.close()

        with zipfile.ZipFile(path) as archive:
            entries = {name: archive.read(name) for name in archive.namelist()}
        sheet_xml = entries["xl/worksheets/sheet1.xml"]
        sheet_xml, replacements = re.subn(
            rb"<f>SUM\(A2:B2\)</f><v>[^<]*</v>",
            b"<f>SUM(A2:B2)</f><v>42</v>",
            sheet_xml,
            count=1,
        )
        self.assertEqual(replacements, 1)
        entries["xl/worksheets/sheet1.xml"] = sheet_xml
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, contents in entries.items():
                archive.writestr(name, contents)

        dataset = load_tabular_file(path, "cached")[0]
        manifest = persist_source(
            self.db,
            "cached",
            path,
            "wiki/sources/cached.md",
            [dataset],
        )[0]

        rows = query_read_only(self.db, f'SELECT * FROM "{manifest.table_name}"')
        schema = query_read_only(
            self.db,
            "SELECT schema_json FROM _openkm_tables WHERE table_name = ?",
            [manifest.table_name],
        )[0]["schema_json"]
        self.assertEqual(rows[0]["a"], 2)
        self.assertEqual(rows[0]["b"], 3)
        self.assertNotIn("total", rows[0])
        self.assertNotIn(42, rows[0].values())
        self.assertNotIn("Total", schema)
        self.assertNotIn("SUM", schema)
        self.assertNotIn("42", schema)

    def test_persistence_catalogs_provenance_samples_and_formula_exclusion(self):
        source = self._source("persist.csv")
        formula = FormulaDefinition(
            column="Calculated",
            formula="=SUM(A2:B2)",
            inputs=("A B", "A-B"),
            sql='coalesce("a_b", 0) + coalesce("a_b_2", 0)',
        )
        first = Dataset(
            name="First Set",
            columns=["A B", "A-B", "Code", "Calculated"],
            rows=[
                {"A B": index, "A-B": index + 1, "Code": f"00{index}", "Calculated": 99}
                for index in range(1, 7)
            ],
            source_page=7,
            source_section="Results",
            source_sheet="Input",
            formulas=[formula],
        )
        second = Dataset(
            name="Second Set",
            columns=["Label"],
            rows=[{"Label": "kept"}],
            source_page=8,
            source_section="Other",
            source_sheet="Other",
        )

        manifests = persist_source(
            self.db,
            "source-1",
            source,
            "wiki/sources/source-1.md",
            [first, second],
        )

        self.assertTrue(self.db.exists())
        self.assertEqual(query_read_only(self.db, "SELECT count(*) AS count FROM _openkm_sources"), [{"count": 1}])
        self.assertEqual(query_read_only(self.db, "SELECT count(*) AS count FROM _openkm_tables"), [{"count": 2}])
        self.assertEqual([manifest.row_count for manifest in manifests], [6, 1])
        self.assertEqual(len(manifests[0].sample), 5)
        self.assertLessEqual(max(len(manifest.sample) for manifest in manifests), 5)

        catalog = query_read_only(
            self.db,
            "SELECT table_name, schema_json FROM _openkm_tables ORDER BY table_name",
        )
        self.assertTrue(all("Calculated" not in row["schema_json"] for row in catalog))
        self.assertTrue(all("SUM" not in row["schema_json"] for row in catalog))

        for manifest, expected_provenance in zip(
            manifests,
            [
                {
                    "source_id": "source-1",
                    "source_page": 7,
                    "source_section": "Results",
                    "source_sheet": "Input",
                },
                {
                    "source_id": "source-1",
                    "source_page": 8,
                    "source_section": "Other",
                    "source_sheet": "Other",
                },
            ],
        ):
            self.assertEqual(
                [name for name, _type in manifest.columns][-4:],
                ["source_id", "source_page", "source_section", "source_sheet"],
            )
            rows = query_read_only(self.db, f'SELECT * FROM "{manifest.table_name}"')
            self.assertTrue(rows)
            self.assertTrue(
                all(
                    row[key] == value
                    for row in rows
                    for key, value in expected_provenance.items()
                )
            )

        first_rows = query_read_only(self.db, f'SELECT * FROM "{manifests[0].table_name}"')
        self.assertEqual(len(first_rows), 6)
        self.assertTrue(all("calculated" not in row for row in first_rows))
        self.assertEqual(first_rows[0]["code"], "001")
        self.assertEqual(
            sorted(row["dataset_name"] for row in query_read_only(
                self.db, "SELECT dataset_name FROM _openkm_tables"
            )),
            ["First Set", "Second Set"],
        )

    def test_persistence_infers_types_preserves_mixed_values_nulls_and_identifiers(self):
        source = self._source("types.csv")
        dataset = Dataset(
            name="Types",
            columns=["Count", "Ratio", "Mixed", "Nothing", "Account ID"],
            rows=[
                {
                    "Count": 1,
                    "Ratio": 1,
                    "Mixed": 1,
                    "Nothing": None,
                    "Account ID": "0012",
                },
                {
                    "Count": 2,
                    "Ratio": 2.5,
                    "Mixed": "two",
                    "Nothing": None,
                    "Account ID": "0013",
                },
            ],
        )

        manifest = persist_source(
            self.db,
            "types",
            source,
            "types.md",
            [dataset],
        )[0]

        self.assertEqual(
            manifest.columns[:5],
            [
                ("count", "BIGINT"),
                ("ratio", "DOUBLE"),
                ("mixed", "VARCHAR"),
                ("nothing", "VARCHAR"),
                ("account_id", "VARCHAR"),
            ],
        )
        rows = query_read_only(
            self.db,
            f'SELECT "count", "ratio", "mixed", "nothing", "account_id" '
            f'FROM "{manifest.table_name}" ORDER BY "count"',
        )
        self.assertEqual(
            rows,
            [
                {
                    "count": 1,
                    "ratio": 1.0,
                    "mixed": "1",
                    "nothing": None,
                    "account_id": "0012",
                },
                {
                    "count": 2,
                    "ratio": 2.5,
                    "mixed": "two",
                    "nothing": None,
                    "account_id": "0013",
                },
            ],
        )

    def test_rebuild_is_idempotent_and_isolates_sources(self):
        one_source = self._source("one.csv", "one")
        two_source = self._source("two.csv", "two")
        one_manifest = persist_source(
            self.db,
            "one",
            one_source,
            "one.md",
            [self._dataset("Items", ["old"])],
        )[0]
        two_manifest = persist_source(
            self.db,
            "two",
            two_source,
            "two.md",
            [self._dataset("Items", ["other"])],
        )[0]

        rebuilt = rebuild_source(
            self.db,
            "one",
            one_source,
            "one-new.md",
            [self._dataset("Items", ["new-1", "new-2"])],
        )[0]
        rebuilt_again = rebuild_source(
            self.db,
            "one",
            one_source,
            "one-new.md",
            [self._dataset("Items", ["new-1", "new-2"])],
        )[0]

        self.assertEqual(rebuilt.table_name, one_manifest.table_name)
        self.assertEqual(rebuilt_again.table_name, one_manifest.table_name)
        self.assertEqual(
            query_read_only(self.db, "SELECT count(*) AS count FROM _openkm_tables"),
            [{"count": 2}],
        )
        self.assertEqual(
            query_read_only(self.db, "SELECT count(*) AS count FROM _openkm_sources"),
            [{"count": 2}],
        )
        self.assertEqual(
            query_read_only(self.db, f'SELECT "value" FROM "{rebuilt.table_name}"'),
            [{"value": "new-1"}, {"value": "new-2"}],
        )
        self.assertEqual(
            query_read_only(self.db, f'SELECT "value" FROM "{two_manifest.table_name}"'),
            [{"value": "other"}],
        )

    def test_failed_persistence_rolls_back_all_source_changes(self):
        stable_source = self._source("stable.csv", "stable")
        broken_source = self._source("broken.csv", "broken")
        stable_manifest = persist_source(
            self.db,
            "stable",
            stable_source,
            "stable.md",
            [self._dataset("Items", ["kept"])],
        )[0]
        good = self._dataset("Good", ["written before failure"])
        bad = Dataset(name="Bad", columns=["Value"], rows=[{"Value": object()}])

        with self.assertRaisesRegex(RuntimeError, "broken"):
            persist_source(
                self.db,
                "broken",
                broken_source,
                "broken.md",
                [good, bad],
            )

        self.assertEqual(
            query_read_only(self.db, "SELECT source_id FROM _openkm_sources"),
            [{"source_id": "stable"}],
        )
        self.assertEqual(
            query_read_only(self.db, "SELECT source_id FROM _openkm_tables"),
            [{"source_id": "stable"}],
        )
        physical_tables = {
            row["name"] for row in query_read_only(self.db, "SHOW TABLES")
        }
        self.assertIn(stable_manifest.table_name, physical_tables)
        self.assertNotIn("broken__good", physical_tables)
        self.assertNotIn("broken__bad", physical_tables)
        self.assertEqual(
            query_read_only(self.db, f'SELECT "value" FROM "{stable_manifest.table_name}"'),
            [{"value": "kept"}],
        )

    def test_query_read_only_supports_select_and_parameters_and_rejects_writes(self):
        source = self._source("query.csv")
        manifest = persist_source(
            self.db,
            "query",
            source,
            "query.md",
            [Dataset(name="Items", columns=["Code"], rows=[{"Code": "0012"}])],
        )[0]

        self.assertEqual(
            query_read_only(
                self.db,
                f'SELECT "code" FROM "{manifest.table_name}" WHERE "code" = ?',
                ["0012"],
            ),
            [{"code": "0012"}],
        )
        self.assertEqual(query_read_only(self.db, "SELECT 1 AS value"), [{"value": 1}])

        for sql in (
            f'INSERT INTO "{manifest.table_name}" VALUES (\'0013\', NULL, NULL, NULL)',
            "CREATE TABLE unwanted (value INTEGER)",
            f'UPDATE "{manifest.table_name}" SET "code" = \'0013\'',
            f'DELETE FROM "{manifest.table_name}"',
            f'DROP TABLE "{manifest.table_name}"',
            f'ALTER TABLE "{manifest.table_name}" RENAME TO unwanted',
            "COPY (SELECT 1) TO 'unwanted.csv'",
            "ATTACH ':memory:' AS unwanted",
            "PRAGMA enable_progress_bar",
            "SELECT 1 -- comment",
            "SELECT 1 /* comment */",
            "SELECT 1; SELECT 2",
        ):
            with self.subTest(sql=sql), self.assertRaises(ValueError):
                query_read_only(self.db, sql)

    def _persist_single(self, source_name, dataset_name, columns, rows):
        source = self._source(source_name)
        return persist_source(
            self.db,
            Path(source_name).stem,
            source,
            f"wiki/sources/{Path(source_name).stem}.md",
            [Dataset(name=dataset_name, columns=columns, rows=rows)],
        )

    def test_normalization_types_and_stores_br_documents_and_dates(self):
        manifests = self._persist_single(
            "pessoas.csv",
            "pessoas",
            ["Nome", "CPF", "Data Nascimento", "Telefone"],
            [
                {"Nome": "Ana", "CPF": "045.106.365-09", "Data Nascimento": "02/09/1990", "Telefone": "(11) 3456-7890"},
                {"Nome": "Beto", "CPF": "12345678901", "Data Nascimento": "31/02/1990", "Telefone": "11 99999-8888"},
            ],
        )
        columns = dict(manifests[0].columns)
        self.assertEqual(columns["cpf"], "VARCHAR")
        self.assertEqual(columns["data_nascimento"], "DATE")
        self.assertEqual(columns["telefone"], "VARCHAR")

        stored = query_read_only(
            self.db,
            'SELECT cpf, data_nascimento, telefone FROM "pessoas__pessoas" ORDER BY cpf',
        )
        self.assertEqual(stored[0]["cpf"], "04510636509")          # leading zero kept, stored VARCHAR
        self.assertEqual(str(stored[0]["data_nascimento"]), "1990-09-02")
        self.assertIsNone(stored[1]["data_nascimento"])             # calendar-invalid -> NULL
        self.assertEqual(stored[0]["telefone"], "551134567890")

        report = {item["column"]: item for item in manifests[0].normalizations}
        self.assertEqual(report["Data Nascimento"]["kind"], "date")
        self.assertEqual(report["Data Nascimento"]["normalized"], 1)
        self.assertEqual(report["Data Nascimento"]["nulled"], 1)    # "31/02/1990" is not a blank token
        self.assertEqual(report["Data Nascimento"]["nulled_examples"], ["31/02/1990"])
        self.assertEqual(report["CPF"]["kind"], "cpf")
        self.assertNotIn("Nome", report)                            # plain text: not normalized

    def test_timestamp_column_is_typed_and_queryable(self):
        manifests = self._persist_single(
            "eventos.csv",
            "eventos",
            ["Ocorrencia", "data_registro"],
            [
                {"Ocorrencia": "a", "data_registro": "02/09/2026 14:30"},
                {"Ocorrencia": "b", "data_registro": "2026-09-02T14:30:00-04:00"},
            ],
        )
        self.assertEqual(dict(manifests[0].columns)["data_registro"], "TIMESTAMP")
        stored = query_read_only(
            self.db,
            "SELECT count(*) AS n FROM eventos__eventos WHERE data_registro >= TIMESTAMP '2026-09-02 18:00:00'",
        )
        self.assertEqual(stored[0]["n"], 1)  # 14:30-04:00 = 18:30 UTC; plain 14:30 stays 14:30
        stored = query_read_only(
            self.db,
            "SELECT count(*) AS n FROM eventos__eventos WHERE data_registro >= TIMESTAMP '2026-09-02 14:00:00'",
        )
        self.assertEqual(stored[0]["n"], 2)

    def test_cpf_digits_only_column_stays_varchar_not_bigint(self):
        # regression: all-digit CPFs used to be inferred BIGINT and lost leading zeros
        manifests = self._persist_single(
            "so_digitos.csv", "t", ["cpf"], [{"cpf": "04510636509"}]
        )
        self.assertEqual(dict(manifests[0].columns)["cpf"], "VARCHAR")
        stored = query_read_only(self.db, 'SELECT cpf FROM "so_digitos__t"')
        self.assertEqual(stored[0]["cpf"], "04510636509")

    def test_free_text_and_mixed_columns_keep_current_behavior(self):
        manifests = self._persist_single(
            "notas.csv",
            "notas",
            ["observacao", "codigo"],
            [
                {"observacao": "02/09/2026", "codigo": "0012"},
                {"observacao": "nada demais", "codigo": "0034"},
            ],
        )
        columns = dict(manifests[0].columns)
        self.assertEqual(columns["observacao"], "VARCHAR")
        self.assertEqual(columns["codigo"], "VARCHAR")  # leading zeros: existing identifier rule
        self.assertEqual(manifests[0].normalizations, [])

    def test_rebuild_of_normalized_dataset_is_idempotent(self):
        columns = ["cpf", "data_nascimento"]
        rows = [{"cpf": "045.106.365-09", "data_nascimento": "02/09/1990"}]
        first = self._persist_single("idem.csv", "idem", columns, rows)
        normalized = query_read_only(self.db, 'SELECT cpf, data_nascimento FROM "idem__idem"')
        again_rows = [
            {"cpf": normalized[0]["cpf"], "data_nascimento": str(normalized[0]["data_nascimento"])}
        ]
        source = self.raw / "idem.csv"
        second = rebuild_source(
            self.db, "idem", source, "wiki/sources/idem.md",
            [Dataset(name="idem", columns=columns, rows=again_rows)],
        )
        self.assertEqual(dict(first[0].columns), dict(second[0].columns))
        after = query_read_only(self.db, 'SELECT cpf, data_nascimento FROM "idem__idem"')
        self.assertEqual(str(after[0]["cpf"]), "04510636509")
        self.assertEqual(str(after[0]["data_nascimento"]), "1990-09-02")


class InspectCliTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        (self.root / "raw").mkdir()
        self.previous_root = os.environ.get("OKM_WIKI_ROOT")
        os.environ["OKM_WIKI_ROOT"] = str(self.root)

    def tearDown(self):
        if self.previous_root is None:
            os.environ.pop("OKM_WIKI_ROOT", None)
        else:
            os.environ["OKM_WIKI_ROOT"] = self.previous_root
        self.temporary_directory.cleanup()

    def test_inspect_file_reports_kind_and_normalized_sample(self):
        path = self.root / "raw" / "cli.csv"
        path.write_text("Nome,CPF,Data Nascimento\nAna,045.106.365-09,02/09/1990\n", encoding="utf-8")
        from scripts.tabular import main

        buffer = io.StringIO()
        with contextlib.redirect_stdout(buffer):
            main(["inspect-file", "--source-id", "cli", "--file", str(path)])
        payload = json.loads(buffer.getvalue())
        columns = {item["name"]: item for item in payload["datasets"][0]["columns"]}
        self.assertEqual(columns["CPF"]["kind"], "cpf")
        self.assertEqual(columns["CPF"]["type"], "VARCHAR")
        self.assertEqual(columns["Data Nascimento"]["kind"], "date")
        self.assertEqual(columns["Data Nascimento"]["type"], "DATE")
        self.assertEqual(payload["datasets"][0]["sample"][0]["CPF"], "04510636509")


if __name__ == "__main__":
    unittest.main()
