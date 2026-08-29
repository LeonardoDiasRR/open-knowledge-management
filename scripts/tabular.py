#!/usr/bin/env python3
"""Load, persist, and query the tabular data owned by OpenKM sources."""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import numbers
import os
import re
import sys
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Sequence

import duckdb


@dataclass(frozen=True)
class FormulaDefinition:
    column: str
    formula: str
    inputs: tuple[str, ...]
    sql: str | None


@dataclass
class Dataset:
    name: str
    columns: list[str]
    rows: list[dict[str, object]]
    source_page: int | None = None
    source_section: str | None = None
    source_sheet: str | None = None
    formulas: list[FormulaDefinition] = field(default_factory=list)


@dataclass(frozen=True)
class TableManifest:
    table_name: str
    dataset_name: str
    row_count: int
    columns: list[tuple[str, str]]
    sample: list[dict[str, object]]


_IDENTIFIER = re.compile(r"^[a-z][a-z0-9_]*$")
_CELL = re.compile(
    r"(?<![A-Za-z0-9_])\$?([A-Za-z]{1,3})\$?(\d+)(?![A-Za-z0-9_])"
)
_REFERENCE = re.compile(
    r"(?<![A-Za-z0-9_])(?:\$?[A-Za-z]{1,3}\$?\d+\s*:\s*\$?[A-Za-z]{1,3}\$?\d+|\$?[A-Za-z]{1,3}\$?\d+)(?![A-Za-z0-9_])"
)
_RANGE = re.compile(
    r"^\$?([A-Za-z]{1,3})\$?(\d+):\$?([A-Za-z]{1,3})\$?(\d+)$"
)
_ARITHMETIC_TOKEN = re.compile(
    r"\s*(?:(?P<cell>\$?[A-Za-z]{1,3}\$?\d+)|"
    r"(?P<number>(?:\d+(?:\.\d*)?|\.\d+))|(?P<operator>[+\-*/()]))"
)
_MUTATING_SQL = re.compile(
    r"\b(?:INSERT|UPDATE|DELETE|CREATE|DROP|ALTER|COPY|ATTACH|INSTALL|LOAD|"
    r"EXPORT|VACUUM|PRAGMA)\b",
    re.IGNORECASE,
)

_PROVENANCE = (
    ("source_id", "VARCHAR"),
    ("source_page", "BIGINT"),
    ("source_section", "VARCHAR"),
    ("source_sheet", "VARCHAR"),
)


def sanitize_identifier(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_value).strip("_").lower()
    slug = slug or "unnamed"
    if slug[0].isdigit():
        slug = "col_" + slug
    return slug


def _unique_identifiers(values: Sequence[str], reserved: set[str] | None = None) -> list[str]:
    used = set(reserved or ())
    result = []
    for value in values:
        base = sanitize_identifier(value)
        candidate = base
        suffix = 2
        while candidate in used:
            candidate = f"{base}_{suffix}"
            suffix += 1
        used.add(candidate)
        result.append(candidate)
    return result


def _quote_identifier(value: str) -> str:
    if not _IDENTIFIER.fullmatch(value):
        raise ValueError(f"unsafe SQL identifier: {value!r}")
    return f'"{value}"'


def _nonempty(value: object) -> bool:
    return value is not None and value != ""


def _read_delimited(path: Path, delimiter: str) -> list[list[str]]:
    text = path.read_text(encoding="utf-8-sig")

    def parse(sep: str) -> list[list[str]]:
        return list(csv.reader(io.StringIO(text), delimiter=sep))

    rows = parse(delimiter)
    nonempty = [row for row in rows if any(_nonempty(cell) for cell in row)]
    if not nonempty:
        return []
    width = len(nonempty[0])
    if all(len(row) == width for row in nonempty):
        return rows

    try:
        sniffed = csv.Sniffer().sniff(text[:8192]).delimiter
    except csv.Error:
        sniffed = delimiter
    if sniffed != delimiter:
        retried = parse(sniffed)
        nonempty = [row for row in retried if any(_nonempty(cell) for cell in row)]
        if nonempty and all(len(row) == len(nonempty[0]) for row in nonempty):
            return retried
    raise ValueError(f"inconsistent row width in {path}")


def _require_unique_headers(headers: Sequence[str], path: Path) -> None:
    seen: dict[str, int] = {}
    for index, header in enumerate(headers, start=1):
        if header in seen:
            raise ValueError(
                f"duplicate header {header!r} in {path} at columns "
                f"{seen[header]} and {index}"
            )
        seen[header] = index


def _load_delimited(path: Path, source_id: str) -> list[Dataset]:
    default = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = _read_delimited(path, default)
    if not rows:
        return []
    nonempty = [row for row in rows if any(_nonempty(cell) for cell in row)]
    headers = nonempty[0]
    _require_unique_headers(headers, path)
    data = []
    for row in nonempty[1:]:
        if len(row) != len(headers):
            raise ValueError(
                f"inconsistent row width in {path}: expected {len(headers)}, got {len(row)}"
            )
        data.append({header: value if value != "" else None for header, value in zip(headers, row)})
    return [Dataset(name=path.stem or source_id, columns=headers, rows=data)]


def _column_number(letters: str) -> int:
    value = 0
    for letter in letters.upper():
        value = value * 26 + ord(letter) - ord("A") + 1
    return value


def _formula_headers(
    formula: str, headers: list[str], physical: list[str | None]
) -> tuple[str, ...]:
    result = []
    for reference in _REFERENCE.finditer(formula):
        range_match = _RANGE.fullmatch(reference.group(0).replace(" ", ""))
        if range_match:
            first = _column_number(range_match.group(1))
            last = _column_number(range_match.group(3))
            indexes = range(first, last + 1) if first <= last else range(last, first + 1)
        else:
            indexes = [_column_number(reference.group(0).replace("$", "").rstrip("0123456789"))]
        for number in indexes:
            index = number - 1
            if (
                0 <= index < len(headers)
                and index < len(physical)
                and physical[index] is not None
            ):
                if headers[index] not in result:
                    result.append(headers[index])
    return tuple(result)


def _formula_cell_sql(
    letters: str,
    headers: list[str],
    physical: list[str | None],
    numeric: list[bool] | None = None,
) -> str | None:
    index = _column_number(letters.upper()) - 1
    if not 0 <= index < len(headers) or index >= len(physical) or physical[index] is None:
        return None
    if numeric is not None and (index >= len(numeric) or not numeric[index]):
        return None
    return _quote_identifier(physical[index])


def _expand_formula_argument(
    argument: str,
    headers: list[str],
    physical: list[str | None],
    formula_row: int | None = None,
    persisted_rows: Sequence[int] | None = None,
    numeric: list[bool] | None = None,
) -> tuple[list[str], bool] | None:
    argument = argument.strip()
    if "$" in argument:
        return None
    match = _RANGE.fullmatch(argument)
    if match:
        start_col, start_row = match.group(1), int(match.group(2))
        end_col, end_row = match.group(3), int(match.group(4))
        if start_row == end_row:
            if formula_row is not None and start_row != formula_row:
                return None
        elif formula_row is not None:
            if (
                persisted_rows is None
                or not any(start_row <= row <= end_row for row in persisted_rows)
                or any(row < start_row or row > end_row for row in persisted_rows)
            ):
                return None
        elif persisted_rows is None or (
            not any(start_row <= row <= end_row for row in persisted_rows)
            or any(row < start_row or row > end_row for row in persisted_rows)
        ):
            return None
        first = _column_number(start_col)
        last = _column_number(end_col)
        if first > last:
            first, last = last, first
        values = []
        for number in range(first, last + 1):
            letters = ""
            current = number
            while current:
                current, remainder = divmod(current - 1, 26)
                letters = chr(ord("A") + remainder) + letters
            sql = _formula_cell_sql(letters, headers, physical, numeric)
            if sql is None:
                return None
            values.append(sql)
        if len(values) != len(set(values)):
            return None
        return values, start_row != end_row

    match = _CELL.fullmatch(argument)
    if not match:
        return None
    if formula_row is not None and int(match.group(2)) != formula_row:
        return None
    sql = _formula_cell_sql(match.group(1), headers, physical, numeric)
    return ([sql], False) if sql else None


def _horizontal_aggregate_sql(name: str, values: list[str]) -> str:
    numeric = [f"try_cast({value} AS DOUBLE)" for value in values]
    total = " + ".join(f"coalesce({value}, 0)" for value in numeric)
    count = " + ".join(f"case when {value} is null then 0 else 1 end" for value in numeric)
    if name == "SUM":
        return f"({total})"
    if name == "AVERAGE":
        return f"(({total}) / nullif(({count}), 0))"

    mean = f"(({total}) / nullif(({count}), 0))"
    squared = " + ".join(
        f"case when {value} is null then 0 else power(({value} - {mean}), 2) end"
        for value in numeric
    )
    divisor = count if name in {"STDEV.P", "STDDEV_POP"} else f"(({count}) - 1)"
    return f"sqrt(({squared}) / nullif({divisor}, 0))"


def _vertical_aggregate_sql(name: str, values: list[str]) -> str:
    numeric = [f"try_cast({value} AS DOUBLE)" for value in values]
    if len(numeric) == 1:
        value = numeric[0]
        if name == "SUM":
            return f"coalesce(sum({value}), 0)"
        if name == "AVERAGE":
            return f"avg({value})"
        function = "stddev_pop" if name in {"STDEV.P", "STDDEV_POP"} else "stddev_samp"
        return f"{function}({value})"

    totals = " + ".join(f"coalesce(sum({value}), 0)" for value in numeric)
    counts = " + ".join(f"count({value})" for value in numeric)
    if name == "SUM":
        return f"({totals})"
    if name == "AVERAGE":
        return f"(({totals}) / nullif(({counts}), 0))"

    squares = " + ".join(f"coalesce(sum(power({value}, 2)), 0)" for value in numeric)
    divisor = counts if name in {"STDEV.P", "STDDEV_POP"} else f"(({counts}) - 1)"
    variance = f"(({squares}) - power(({totals}), 2) / nullif(({counts}), 0))"
    return f"sqrt(({variance}) / nullif({divisor}, 0))"


def _translate_formula(
    formula: str,
    headers: list[str],
    physical: list[str | None],
    formula_row: int | None = None,
    persisted_rows: Sequence[int] | None = None,
    numeric: list[bool] | None = None,
) -> str | None:
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    body = formula[1:].strip()
    function = re.fullmatch(
        r"(?i)(SUM|AVERAGE|STDEV(?:\.S|\.P)?|STDDEV(?:_SAMP|_POP)?)\s*\((.*)\)",
        body,
    )
    if function:
        name = function.group(1).upper()
        arguments = function.group(2).split(",")
        if not arguments or any(not part.strip() for part in arguments):
            return None
        values = []
        vertical = None
        for argument in arguments:
            expanded = _expand_formula_argument(
                argument,
                headers,
                physical,
                formula_row,
                persisted_rows,
                numeric,
            )
            if expanded is None:
                return None
            argument_values, argument_vertical = expanded
            if vertical is None:
                vertical = argument_vertical
            elif vertical != argument_vertical:
                return None
            values.extend(argument_values)
        if not values:
            return None
        if len(values) != len(set(values)):
            return None
        aggregate = _vertical_aggregate_sql if vertical else _horizontal_aggregate_sql
        return aggregate(name, values)

    references = list(_CELL.finditer(body))
    if any("$" in match.group(0) for match in references):
        return None
    rows = {int(match.group(2)) for match in references}
    if len(rows) > 1 or (formula_row is not None and any(row != formula_row for row in rows)):
        return None

    tokens = []
    position = 0
    while position < len(body):
        match = _ARITHMETIC_TOKEN.match(body, position)
        if match is None:
            return None
        tokens.append(match)
        position = match.end()

    token_index = 0

    def parse_factor() -> str | None:
        nonlocal token_index
        if token_index >= len(tokens):
            return None
        token = tokens[token_index]
        operator = token.group("operator")
        if operator in {"+", "-"}:
            token_index += 1
            value = parse_factor()
            return f"{operator}{value}" if value is not None else None
        if operator == "(":
            token_index += 1
            value = parse_expression()
            if value is None or token_index >= len(tokens) or tokens[token_index].group("operator") != ")":
                return None
            token_index += 1
            return f"({value})"
        if token.group("cell"):
            cell_match = _CELL.fullmatch(token.group("cell"))
            cell_sql = _formula_cell_sql(cell_match.group(1), headers, physical, numeric)
            if cell_sql is None:
                return None
            token_index += 1
            return f"coalesce(try_cast({cell_sql} AS DOUBLE), 0)"
        if token.group("number"):
            token_index += 1
            return token.group("number")
        return None

    def parse_term() -> str | None:
        nonlocal token_index
        value = parse_factor()
        while value is not None and token_index < len(tokens):
            operator = tokens[token_index].group("operator")
            if operator not in {"*", "/"}:
                break
            token_index += 1
            right = parse_factor()
            if right is None:
                return None
            if operator == "/":
                right = f"nullif({right}, 0)"
            value = f"{value}{operator}{right}"
        return value

    def parse_expression() -> str | None:
        nonlocal token_index
        value = parse_term()
        while value is not None and token_index < len(tokens):
            operator = tokens[token_index].group("operator")
            if operator not in {"+", "-"}:
                break
            token_index += 1
            right = parse_term()
            if right is None:
                return None
            value = f"{value}{operator}{right}"
        return value

    translated_sql = parse_expression()
    if translated_sql is None or token_index != len(tokens):
        return None
    return None if "--" in translated_sql else translated_sql


def _load_xlsx(path: Path) -> list[Dataset]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read XLSX files") from exc

    workbook = load_workbook(path, read_only=False, data_only=False)
    datasets = []
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows())
            header_index = next(
                (index for index, row in enumerate(rows) if any(_nonempty(cell.value) for cell in row)),
                None,
            )
            if header_index is None:
                continue
            header_cells = rows[header_index]
            headers = ["" if cell.value is None else str(cell.value) for cell in header_cells]
            _require_unique_headers(headers, path)
            all_data_rows = [
                (row_number, row)
                for row_number, row in enumerate(
                    rows[header_index + 1 :], start=header_index + 2
                )
                if any(_nonempty(cell.value) for cell in row)
            ]
            if not all_data_rows:
                continue
            formula_columns = {
                index
                for _row_number, row in all_data_rows
                for index, cell in enumerate(row)
                if _nonempty(cell.value) and cell.data_type == "f"
            }
            data_rows = [
                (row_number, row)
                for row_number, row in all_data_rows
                if any(
                    index not in formula_columns
                    and index < len(row)
                    and _nonempty(row[index].value)
                    for index in range(len(headers))
                )
            ]
            if not data_rows:
                continue
            physical = _unique_identifiers(
                headers, reserved={name for name, _type in _PROVENANCE}
            )
            formula_physical = [
                None if index in formula_columns else name
                for index, name in enumerate(physical)
            ]
            numeric_inputs = []
            for index in range(len(headers)):
                values = [
                    row[index].value
                    for _row_number, row in data_rows
                    if index < len(row) and _nonempty(row[index].value)
                ]
                numeric_inputs.append(
                    bool(values)
                    and all(
                        isinstance(value, numbers.Number) and not isinstance(value, bool)
                        for value in values
                    )
                )
            persisted_row_numbers = [
                row_number
                for row_number, row in data_rows
                if any(
                    index not in formula_columns
                    and index < len(row)
                    and _nonempty(row[index].value)
                    for index in range(len(headers))
                )
            ]
            persisted_rows = tuple(persisted_row_numbers) or None
            formulas = []
            for index in sorted(formula_columns):
                formula_cells = [
                    (row_number, str(row[index].value))
                    for row_number, row in all_data_rows
                    if index < len(row) and row[index].data_type == "f"
                ]
                definitions = {}
                for row_number, value in formula_cells:
                    sql = _translate_formula(
                        value,
                        headers,
                        formula_physical,
                        formula_row=row_number,
                        persisted_rows=persisted_rows,
                        numeric=numeric_inputs,
                    )
                    key = (value, sql)
                    if key not in definitions:
                        definitions[key] = [value, sql, []]
                    definitions[key][2].append(value)
                for formula, sql, values in definitions.values():
                    inputs = []
                    for value in values:
                        for header in _formula_headers(value, headers, formula_physical):
                            if header not in inputs:
                                inputs.append(header)
                    formulas.append(
                        FormulaDefinition(
                            column=headers[index] if index < len(headers) else "",
                            formula=formula,
                            inputs=tuple(inputs),
                            sql=sql,
                        )
                    )
            columns = [header for index, header in enumerate(headers) if index not in formula_columns]
            records = []
            for _row_number, row in data_rows:
                record = {}
                for index, header in enumerate(headers):
                    if index not in formula_columns:
                        value = row[index].value if index < len(row) else None
                        record[header] = None if value == "" else value
                records.append(record)
            datasets.append(
                Dataset(
                    name=worksheet.title,
                    columns=columns,
                    rows=records,
                    source_sheet=worksheet.title,
                    formulas=formulas,
                )
            )
    finally:
        workbook.close()
    return datasets


def load_tabular_file(path: Path, source_id: str) -> list[Dataset]:
    path = _require_raw_file(Path(path))
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return _load_delimited(path, source_id)
    if suffix == ".xlsx":
        return _load_xlsx(path)
    raise ValueError(f"unsupported tabular file extension: {path.suffix or '<none>'}")


def _numeric_string(value: str) -> tuple[str, int | float] | None:
    if re.fullmatch(r"[-+]?(?:0|[1-9]\d*)", value):
        return "BIGINT", int(value)
    if re.fullmatch(r"[-+]?(?:0|[1-9]\d*)\.\d+", value):
        return "DOUBLE", float(value)
    return None


def _iso_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None


def _infer_type(values: Sequence[object]) -> str:
    nonnull = [value for value in values if value is not None]
    if not nonnull:
        return "VARCHAR"
    if all(isinstance(value, bool) for value in nonnull):
        return "BOOLEAN"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in nonnull):
        return "BIGINT"
    if all(isinstance(value, numbers.Number) and not isinstance(value, bool) for value in nonnull):
        return "DOUBLE"
    if all(isinstance(value, str) for value in nonnull):
        numeric = [_numeric_string(value) for value in nonnull]
        if all(item is not None for item in numeric):
            return "BIGINT" if all(item[0] == "BIGINT" for item in numeric) else "DOUBLE"
    if all(_iso_date(value) is not None for value in nonnull):
        return "DATE"
    return "VARCHAR"


def _typed_value(value: object, type_name: str) -> object:
    if value is None:
        return None
    if type_name == "DATE":
        return _iso_date(value)
    if type_name == "BIGINT" and isinstance(value, str):
        return int(value)
    if type_name == "DOUBLE" and isinstance(value, str):
        return float(value)
    return value


def _require_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"source file does not exist: {path}")
    if not path.is_file():
        raise ValueError(f"source path is not a file: {path}")


def _configured_raw_dir() -> Path:
    configured_root = os.environ.get("OKM_WIKI_ROOT")
    if not configured_root:
        raise ValueError(
            "OKM_WIKI_ROOT must be set to the wiki root before accessing raw sources"
        )
    return (Path(configured_root).expanduser().resolve() / "raw").resolve()


def _require_raw_file(path: Path) -> Path:
    candidate = Path(path).expanduser().resolve()
    configured_raw = _configured_raw_dir().resolve()
    try:
        candidate.relative_to(configured_raw)
    except ValueError as exc:
        raise ValueError(
            f"source path must be inside the configured wiki raw directory: {candidate}"
        ) from exc
    _require_file(candidate)
    return candidate


def _checksum(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _catalog_sql(connection: duckdb.DuckDBPyConnection) -> None:
    connection.execute(
        """CREATE TABLE IF NOT EXISTS _openkm_sources (
            source_id VARCHAR PRIMARY KEY,
            source_resource VARCHAR NOT NULL,
            summary_page VARCHAR NOT NULL,
            content_checksum VARCHAR NOT NULL,
            ingested_at TIMESTAMP NOT NULL,
            extraction_status VARCHAR NOT NULL
        )"""
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS _openkm_tables (
            table_name VARCHAR PRIMARY KEY,
            source_id VARCHAR NOT NULL,
            dataset_name VARCHAR NOT NULL,
            schema_json VARCHAR NOT NULL,
            row_count BIGINT NOT NULL,
            updated_at TIMESTAMP NOT NULL
        )"""
    )


def _table_names(connection: duckdb.DuckDBPyConnection) -> set[str]:
    return {row[0] for row in connection.execute("SELECT table_name FROM _openkm_tables").fetchall()}


def _persist(
    db_path: Path,
    source_id: str,
    source_resource: Path,
    summary_page: str,
    datasets: Sequence[Dataset],
    rebuild: bool,
) -> list[TableManifest]:
    if not source_id:
        raise ValueError("source_id must not be empty")
    db_path = Path(db_path)
    source_resource = _require_raw_file(Path(source_resource))
    db_path.parent.mkdir(parents=True, exist_ok=True)
    connection = duckdb.connect(str(db_path))
    try:
        connection.execute("BEGIN")
        _catalog_sql(connection)
        existing = connection.execute(
            "SELECT 1 FROM _openkm_sources WHERE source_id = ?", [source_id]
        ).fetchone()
        if existing and not rebuild:
            raise ValueError(
                f"source {source_id!r} already exists; call rebuild_source to replace it"
            )
        if rebuild and not existing:
            raise ValueError(f"source {source_id!r} does not exist; call persist_source first")

        if rebuild:
            owned = connection.execute(
                "SELECT table_name FROM _openkm_tables WHERE source_id = ?", [source_id]
            ).fetchall()
            for (table_name,) in owned:
                connection.execute(f"DROP TABLE IF EXISTS {_quote_identifier(table_name)}")
            connection.execute("DELETE FROM _openkm_tables WHERE source_id = ?", [source_id])
            connection.execute("DELETE FROM _openkm_sources WHERE source_id = ?", [source_id])

        used_tables = _table_names(connection)
        source_slug = sanitize_identifier(source_id)
        manifests = []
        for dataset in datasets:
            dataset_slug = sanitize_identifier(dataset.name)
            base_table = f"{source_slug}__{dataset_slug}"
            table_name = base_table
            suffix = 2
            while table_name in used_tables:
                table_name = f"{base_table}_{suffix}"
                suffix += 1
            used_tables.add(table_name)

            formula_columns = {formula.column for formula in dataset.formulas}
            stored_columns = [
                column for column in dataset.columns if column not in formula_columns
            ]
            physical_columns = _unique_identifiers(
                stored_columns, reserved={name for name, _type in _PROVENANCE}
            )
            column_types = [
                _infer_type([row.get(column) for row in dataset.rows])
                for column in stored_columns
            ]
            schema = list(zip(physical_columns, column_types)) + list(_PROVENANCE)
            definitions = ", ".join(
                f"{_quote_identifier(name)} {type_name}" for name, type_name in schema
            )
            connection.execute(f"CREATE TABLE {_quote_identifier(table_name)} ({definitions})")

            all_columns = physical_columns + [name for name, _type in _PROVENANCE]
            quoted_columns = ", ".join(_quote_identifier(name) for name in all_columns)
            placeholders = ", ".join("?" for _ in all_columns)
            values = []
            for row in dataset.rows:
                values.append(
                    [
                        _typed_value(row.get(original), type_name)
                        for original, type_name in zip(stored_columns, column_types)
                    ]
                    + [
                        source_id,
                        dataset.source_page,
                        dataset.source_section,
                        dataset.source_sheet,
                    ]
                )
            if values:
                connection.executemany(
                    f"INSERT INTO {_quote_identifier(table_name)} ({quoted_columns}) VALUES ({placeholders})",
                    values,
                )
            schema_json = json.dumps(
                [{"name": name, "type": type_name} for name, type_name in schema],
                ensure_ascii=True,
            )
            connection.execute(
                """INSERT INTO _openkm_tables
                   (table_name, source_id, dataset_name, schema_json, row_count, updated_at)
                   VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)""",
                [table_name, source_id, dataset.name, schema_json, len(dataset.rows)],
            )
            manifests.append(
                TableManifest(
                    table_name=table_name,
                    dataset_name=dataset.name,
                    row_count=len(dataset.rows),
                    columns=schema,
                    sample=[
                        {
                            name: _typed_value(row.get(original), type_name)
                            for original, name, type_name in zip(
                                stored_columns, physical_columns, column_types
                            )
                        }
                        | {
                            "source_id": source_id,
                            "source_page": dataset.source_page,
                            "source_section": dataset.source_section,
                            "source_sheet": dataset.source_sheet,
                        }
                        for row in dataset.rows[:5]
                    ],
                )
            )

        connection.execute(
            """INSERT INTO _openkm_sources
               (source_id, source_resource, summary_page, content_checksum,
                ingested_at, extraction_status)
               VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP, ?)""",
            [source_id, str(source_resource), summary_page, _checksum(Path(source_resource)), "complete"],
        )
        connection.execute("COMMIT")
        return manifests
    except Exception as exc:
        try:
            connection.execute("ROLLBACK")
        except Exception:
            pass
        if isinstance(exc, ValueError) and (
            "already exists; call rebuild_source" in str(exc)
            or "does not exist; call persist_source" in str(exc)
        ):
            raise
        raise RuntimeError(f"failed to persist source {source_id!r}: {exc}") from exc
    finally:
        connection.close()


def persist_source(
    db_path: Path,
    source_id: str,
    source_resource: Path,
    summary_page: str,
    datasets: Sequence[Dataset],
) -> list[TableManifest]:
    return _persist(db_path, source_id, source_resource, summary_page, datasets, rebuild=False)


def rebuild_source(
    db_path: Path,
    source_id: str,
    source_resource: Path,
    summary_page: str,
    datasets: Sequence[Dataset],
) -> list[TableManifest]:
    return _persist(db_path, source_id, source_resource, summary_page, datasets, rebuild=True)


def _sql_without_literals(sql: str) -> tuple[str, int, bool]:
    output = []
    semicolons = 0
    index = 0
    while index < len(sql):
        char = sql[index]
        if char in "'\"`":
            quote = char
            output.append(" ")
            index += 1
            while index < len(sql):
                if sql[index] == quote:
                    if index + 1 < len(sql) and sql[index + 1] == quote:
                        index += 2
                        continue
                    index += 1
                    break
                index += 1
            continue
        if sql.startswith("--", index):
            newline = sql.find("\n", index)
            index = len(sql) if newline < 0 else newline
            output.append("--")
            continue
        if sql.startswith("/*", index):
            end = sql.find("*/", index + 2)
            if end < 0:
                return "", 0, False
            index = end + 2
            output.append("/*")
            continue
        if char == ";":
            semicolons += 1
        output.append(char)
        index += 1
    return "".join(output), semicolons, True


def query_read_only(
    db_path: Path,
    sql: str,
    parameters: Sequence[object] = (),
) -> list[dict[str, object]]:
    cleaned, semicolons, valid = _sql_without_literals(sql)
    if not valid:
        raise ValueError("invalid SQL comments or quoted text")
    if "--" in cleaned or "/*" in cleaned:
        raise ValueError("SQL comments are not allowed")
    if semicolons and cleaned.rstrip().endswith(";"):
        cleaned = cleaned.rstrip()[:-1].rstrip()
    if ";" in cleaned or semicolons > 1:
        raise ValueError("multiple SQL statements are not allowed")
    if _MUTATING_SQL.search(cleaned):
        raise ValueError("write SQL is not allowed")
    if not re.match(r"(?is)^\s*(?:SELECT|WITH|DESCRIBE|SHOW)\b", cleaned):
        raise ValueError("only SELECT, read-only WITH, DESCRIBE, and SHOW are allowed")

    connection = duckdb.connect(str(Path(db_path)), read_only=True)
    try:
        cursor = connection.execute(sql, list(parameters))
        columns = [description[0] for description in cursor.description or ()]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]
    finally:
        connection.close()


def _formula_from_json(value: object) -> FormulaDefinition:
    if not isinstance(value, dict):
        raise ValueError("formula definitions must be JSON objects")
    return FormulaDefinition(
        column=str(value["column"]),
        formula=str(value["formula"]),
        inputs=tuple(str(item) for item in value.get("inputs", ())),
        sql=value.get("sql"),
    )


def _datasets_from_json(payload: dict[str, object]) -> list[Dataset]:
    values = payload.get("datasets")
    if not isinstance(values, list):
        raise ValueError("input JSON must contain a datasets array")
    datasets = []
    for value in values:
        if not isinstance(value, dict):
            raise ValueError("each dataset must be a JSON object")
        datasets.append(
            Dataset(
                name=str(value["name"]),
                columns=[str(column) for column in value["columns"]],
                rows=[dict(row) for row in value["rows"]],
                source_page=value.get("source_page"),
                source_section=value.get("source_section"),
                source_sheet=value.get("source_sheet"),
                formulas=[_formula_from_json(item) for item in value.get("formulas", ())],
            )
        )
    return datasets


def _inspect_payload(source_id: str, datasets: Sequence[Dataset]) -> dict[str, object]:
    result = []
    for dataset in datasets:
        result.append(
            {
                "name": dataset.name,
                "columns": [
                    {"name": column, "type": _infer_type([row.get(column) for row in dataset.rows])}
                    for column in dataset.columns
                ],
                "row_count": len(dataset.rows),
                "sample": dataset.rows[:5],
                "source_page": dataset.source_page,
                "source_section": dataset.source_section,
                "source_sheet": dataset.source_sheet,
                "formulas": [asdict(formula) for formula in dataset.formulas],
            }
        )
    return {"source_id": source_id, "datasets": result}


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="OpenKM DuckDB tabular helper")
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect = subparsers.add_parser("inspect-file")
    inspect.add_argument("--source-id", required=True)
    inspect.add_argument("--file", required=True, type=Path)

    for command in ("persist-json", "rebuild-json"):
        command_parser = subparsers.add_parser(command)
        command_parser.add_argument("--db", required=True, type=Path)
        command_parser.add_argument("--source-id", required=True)
        command_parser.add_argument("--source-resource", required=True, type=Path)
        command_parser.add_argument("--summary-page", required=True)
        command_parser.add_argument("--input", required=True, type=Path)

    query = subparsers.add_parser("query")
    query.add_argument("--db", required=True, type=Path)
    query.add_argument("--sql", required=True)
    query.add_argument("--parameter", action="append", nargs="+", default=[])
    return parser


def _parameter(value: str) -> object:
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return value


def main(argv: Sequence[str] | None = None) -> None:
    args = _build_parser().parse_args(argv)
    if args.command == "inspect-file":
        datasets = load_tabular_file(args.file, args.source_id)
        print(json.dumps(_inspect_payload(args.source_id, datasets), ensure_ascii=True, default=str))
        return
    if args.command == "query":
        parameters = [value for group in args.parameter for value in group]
        print(
            json.dumps(
                query_read_only(args.db, args.sql, [_parameter(value) for value in parameters]),
                ensure_ascii=True,
                default=str,
            )
        )
        return

    _require_file(args.input)
    payload = json.loads(args.input.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or payload.get("source_id") != args.source_id:
        raise ValueError("input JSON source_id must match --source-id")
    datasets = _datasets_from_json(payload)
    persist = persist_source if args.command == "persist-json" else rebuild_source
    manifests = persist(
        args.db,
        args.source_id,
        args.source_resource,
        args.summary_page,
        datasets,
    )
    print(json.dumps({"source_id": args.source_id, "tables": [asdict(item) for item in manifests]}, default=str))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(1)
